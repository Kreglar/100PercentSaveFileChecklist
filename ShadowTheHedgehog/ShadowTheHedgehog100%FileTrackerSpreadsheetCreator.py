# ShadowTheHedgehog100%FileTrackerSpreadheetCreator.py
# Created by Kreglar (Miles Cooper 2025)
# Version 1.0.0

# openpyxl used to create the spreadsheet
import openpyxl as excel
import openpyxl.styles as style

# to get data about the game
import json

def apply_font(sheet, range: str, font: style.Font) -> None:
    """ Apply a font to a range of cells. """
    for row in sheet[range]:
        for cell in row:
            cell.font = font

# retrieve json data
with open("ShadowTheHedgehog.json", 'r') as file:
    json_data = json.load(file)
sequences = json_data["sequences"] # each sequence in the json
stages = json_data["stages"][0] + json_data["stages"][1] + json_data["stages"][2] + json_data["stages"][3] + json_data["stages"][4] + json_data["stages"][5] # each stage in the json

# create the main workbook (container holding spreadsheets)
spreadsheet = excel.Workbook()

# create the sheets for each item in 100%
sequences_sheet = spreadsheet.create_sheet("Sequences") # for all 326 possible story paths
ARank_sheet = spreadsheet.create_sheet("A-Rank") # for an 'A' rank on every ending of every level + boss fights
keys_sheet = spreadsheet.create_sheet("Keys") # for the 5 collectable keys per level

spreadsheet.remove(spreadsheet.active) # delete the default sheet

# create fonts
title_bar_font = style.Font(bold=True)
secondary_font = style.Font(bold=False)

# SEQUENCES SHEET ---------------------------------------------------------------------------------------------------------
# create the top title bar
sequences_sheet["A1"] = "Route Name" # name and number of the sequence
sequences_sheet["B1"] = "Route Code" # which endings to take for each sequence (D = Dark, N = Neutral, H = Hero)
sequences_sheet["C1"] = "Completed?" # label yes or no if completed
sequences_sheet.freeze_panes = "A2" # freeze title row

# set collumn widths
sequences_sheet.column_dimensions["A"].width = 35
sequences_sheet.column_dimensions["B"].width = 15
sequences_sheet.column_dimensions["C"].width = 15

# list out each sequence
for route_number in range(len(sequences)):
    sequences_sheet["A" + str(route_number + 2)] = sequences[route_number]["name"] # input route name
    sequences_sheet["B" + str(route_number + 2)] = sequences[route_number]["route"] # input route code

# apply fonts
apply_font(sequences_sheet, "A1:C1", title_bar_font)
apply_font(sequences_sheet, "A2:C327", secondary_font)

# A-RANK SHEET ------------------------------------------------------------------------------------------------------------
# create the top title bar
ARank_sheet["A1"] = "Stage Name" # name of the current stage
ARank_sheet["B1"] = "Dark Mission" # A-Rank for each mission (yes or no)
ARank_sheet["C1"] = "Neutral Mission"
ARank_sheet["D1"] = "Hero Mission"
ARank_sheet["F1"] = "Boss Name" # name of the current boss
ARank_sheet["G1"] = "Encounter" # name of the level before the encounter
ARank_sheet["H1"] = "A-Rank?" # A-Rank for each boss (yes or no)
ARank_sheet.freeze_panes = "A2"

# set collumn widths
ARank_sheet.column_dimensions["A"].width = 20
ARank_sheet.column_dimensions["B"].width = 20
ARank_sheet.column_dimensions["C"].width = 20
ARank_sheet.column_dimensions["D"].width = 20
ARank_sheet.column_dimensions["F"].width = 20
ARank_sheet.column_dimensions["G"].width = 20

# list out each stage
for level_number in range(len(stages)):
    ARank_sheet["A" + str(level_number + 2)] = stages[level_number]["name"] # input stage name

# list out each boss
bosses = json_data["bosses"]
battle_number = 0
for boss_number in range(len(bosses)):
    encounters = bosses[boss_number]["encounters"] # list out encounters per boss
    for fight_number in range(len(encounters)):
        ARank_sheet["F" + str(battle_number + 2)] = bosses[boss_number]["name"] # label boss name
        ARank_sheet["G" + str(battle_number + 2)] = encounters[fight_number] # label boss encounter
        battle_number += 1

# apply fonts
apply_font(ARank_sheet, "A1:H1", title_bar_font)
apply_font(ARank_sheet, "A2:H23", secondary_font)

# KEYS SHEET --------------------------------------------------------------------------------------------------------------
# create the top title bar
keys_sheet["A1"] = "Stage Name" # name of the current stage
keys_sheet["B1"] = "5 Keys?" # have 5 keys for each stage (yes or no)
keys_sheet.freeze_panes = "A2"

# set collumn widths
keys_sheet.column_dimensions["A"].width = 20

# list out each stage
for level_number in range(len(stages)):
    keys_sheet["A" + str(level_number + 2)] = stages[level_number]["name"] # input stage name

# apply fonts
apply_font(keys_sheet, "A1:B1", title_bar_font)
apply_font(keys_sheet, "A2:B23", secondary_font)

# SAVE THE SHEET ----------------------------------------------------------------------------------------------------------
spreadsheet.save("ShadowTheHedgehog100%.xlsx")
